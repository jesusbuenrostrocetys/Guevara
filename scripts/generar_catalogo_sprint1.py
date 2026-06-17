# -*- coding: utf-8 -*-
"""
SYSEA Radiadores - SPRINT 1: Catálogo Maestro
==============================================

Genera el archivo ENTREGAS/SYSEA_Radiadores_SPRINT1.xlsx con:
  - Hoja "Catálogo Maestro" con las 21 columnas oficiales (A..U)
  - 20 productos ficticios de validación
  - Fórmulas automáticas (precios, utilidades, rotación, alerta AutoZone)
  - Validaciones de datos (DPI único, precios y stock no negativos, dropdown Estado)
  - Formato condicional (Stock, Rotación, Utilidad Cliente)

NOTA TÉCNICA IMPORTANTE
-----------------------
Las fórmulas se guardan en el archivo .xlsx SIEMPRE en inglés y con coma como
separador de argumentos (IF, ROUNDUP, TODAY, MAX). Así lo exige el formato
OOXML. Excel las MOSTRARÁ traducidas al idioma de la interfaz: en español verás
SI, REDONDEAR.MAS, HOY, MAX y separadas por punto y coma (;). El comportamiento
es idéntico; solo cambia la presentación.

CORRECCIÓN DE REFERENCIAS
-------------------------
En el prompt original las fórmulas usaban "C2" para Precio Compra, pero en la
estructura oficial de 21 columnas el Precio Compra está en la COLUMNA F. Se
remapeó C2 -> F2 para que las fórmulas apunten a la celda correcta. La lógica /
política de precios NO se modificó.
"""

import math
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Parámetros del negocio (Sprint 1)                                           #
# --------------------------------------------------------------------------- #
HOY = date(2026, 6, 17)            # fecha de referencia del Sprint 1
STOCK_MINIMO = 1
STOCK_IDEAL = 3
PROVEEDOR = "Radiadores del Centro"

# --------------------------------------------------------------------------- #
# Datos ficticios: 20 productos (DPI, Descripción, Marca, Modelo, Año,        #
#                  Precio Compra, Precio AutoZone)                            #
# --------------------------------------------------------------------------- #
PRODUCTOS = [
    ("RVVW029",   "VW Jetta Radiador",          "Volkswagen",    "Jetta",   2010, 1500, 3900),
    ("0050",      "Nissan Sentra Radiador",     "Nissan",        "Sentra",  2015, 1200, 3200),
    ("13015",     "Chevy Aveo Radiador",        "Chevrolet",     "Aveo",    2012,  800, 2500),
    ("FORD021",   "Ford Focus Radiador",        "Ford",          "Focus",   2013, 1350, 3600),
    ("HYUN008",   "Hyundai Elantra Radiador",   "Hyundai",       "Elantra", 2011,  950, 2800),
    ("KIA045",    "Kia Cerato Radiador",        "Kia",           "Cerato",  2014, 1100, 3100),
    ("MAZU33",    "Mazda 3 Radiador",           "Mazda",         "Mazda 3", 2016, 1600, 4200),
    ("MERC77",    "Mercedes C200 Radiador",     "Mercedes-Benz", "C200",    2012, 3500, 8900),
    ("SUBR12",    "Subaru Impreza Radiador",    "Subaru",        "Impreza", 2010, 1450, 3700),
    ("TOYO88",    "Toyota Corolla Radiador",    "Toyota",        "Corolla", 2015, 1300, 3400),
    ("MITSU19",   "Mitsubishi Lancer Radiador", "Mitsubishi",    "Lancer",  2011, 1050, 2900),
    ("VOLKS30",   "VW Passat Radiador",         "Volkswagen",    "Passat",  2009, 2000, 5200),
    ("CHRY55",    "Chrysler Neon Radiador",     "Chrysler",      "Neon",    2008,  700, 2200),
    ("HONDA64",   "Honda Civic Radiador",       "Honda",         "Civic",   2012, 1250, 3300),
    ("SUZUK16",   "Suzuki Swift Radiador",      "Suzuki",        "Swift",   2014,  900, 2600),
    ("BUICK22",   "Buick Century Radiador",     "Buick",         "Century", 2005,  600, 1900),
    ("JEEP99",    "Jeep Wrangler Radiador",     "Jeep",          "Wrangler",2013, 1800, 4800),
    ("LEXUS50",   "Lexus RX350 Radiador",       "Lexus",         "RX350",   2010, 4000, 9500),
    ("PEUGEOT11", "Peugeot 206 Radiador",       "Peugeot",       "206",     2006,  750, 2400),
    ("RENAULT88", "Renault Duster Radiador",    "Renault",       "Duster",  2015, 1200, 3300),
]

# Stock actual por producto (0..4) -> demuestra rojo/amarillo/verde
STOCK_ACTUAL = [3, 2, 0, 1, 4, 0, 3, 1, 2, 4, 0, 2, 1, 3, 2, 0, 4, 1, 2, 3]

# Días desde la última venta (None = nunca vendido) -> demuestra cada rotación
DIAS_ULTIMA_VENTA = [15, 45, 200, None, 120, 400, 30, 250, 160, 75,
                     500, None, 350, 60, 100, 600, 20, 280, 170, 90]

# Días desde la última compra (1..30)
DIAS_ULTIMA_COMPRA = [5, 12, 20, 3, 28, 15, 8, 22, 10, 18,
                      25, 6, 14, 30, 9, 17, 4, 21, 11, 27]

# Observaciones (solo notas relevantes; el resto en blanco)
OBSERVACIONES = {
    2:  "Reabastecer pronto",
    3:  "Recién agregado al catálogo",
    5:  "Sin venta >1 año, evaluar descuento",
    10: "Sin movimiento, revisar",
    11: "Producto nuevo en catálogo",
    15: "Modelo antiguo, baja demanda",
}

# --------------------------------------------------------------------------- #
# Encabezados (orden EXACTO de las 21 columnas A..U)                          #
# --------------------------------------------------------------------------- #
ENCABEZADOS = [
    "DPI",                  # A
    "Descripción",          # B
    "Marca",                # C
    "Modelo",               # D
    "Año",                  # E
    "Precio Compra",        # F
    "Precio AutoZone",      # G
    "Precio Cliente Final", # H  (fórmula)
    "Precio Taller",        # I  (fórmula)
    "Utilidad Cliente",     # J  (fórmula)
    "Utilidad Taller",      # K  (fórmula)
    "Stock Actual",         # L
    "Stock Mínimo",         # M
    "Stock Ideal",          # N
    "Proveedor Principal",  # O
    "Fecha Última Compra",  # P
    "Fecha Última Venta",   # Q
    "Estado",               # R  (dropdown)
    "Rotación",             # S  (fórmula)
    "Alerta AutoZone",      # T  (fórmula)
    "Observaciones",        # U
]

# --------------------------------------------------------------------------- #
# Estilos                                                                     #
# --------------------------------------------------------------------------- #
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
FORMULA_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")  # columnas calculadas
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Colores de formato condicional (estilo estándar de Excel)
F_VERDE = PatternFill("solid", fgColor="C6EFCE");  T_VERDE = Font(color="006100")
F_AMARI = PatternFill("solid", fgColor="FFEB9C");  T_AMARI = Font(color="9C6500")
F_NARAN = PatternFill("solid", fgColor="FFCC99");  T_NARAN = Font(color="974706")
F_ROJO  = PatternFill("solid", fgColor="FFC7CE");  T_ROJO  = Font(color="9C0006")
F_GRIS  = PatternFill("solid", fgColor="D9D9D9");  T_GRIS  = Font(color="3F3F3F")

MXN = '"$"#,##0'
FECHA = "dd/mm/yyyy"

# Columnas calculadas por fórmula (para colorear el encabezado distinto)
COLS_FORMULA = {"H", "I", "J", "K", "S", "T"}

# --------------------------------------------------------------------------- #
# Helper de verificación (réplica en Python de la lógica de Excel)            #
# --------------------------------------------------------------------------- #
def roundup50(x: float) -> int:
    """Redondea SIEMPRE hacia arriba al múltiplo de 50 más cercano."""
    return int(math.ceil(x / 50.0)) * 50


def precio_cliente(compra: float, autozone: float) -> int:
    if autozone and autozone > 0:
        # quedarse $250 debajo de AutoZone, garantizando utilidad mínima $1,200
        interno = (compra + 1200) if (autozone - compra - 250) < 1200 else (autozone - 250)
        return roundup50(interno)
    return roundup50(compra + 1500)


def precio_taller(compra: float, cliente: float) -> int:
    # descuento ~18% sobre cliente, pero nunca menos de $800 de utilidad
    return roundup50(max(compra + 800, cliente * 0.82))


# --------------------------------------------------------------------------- #
# Construcción del libro                                                      #
# --------------------------------------------------------------------------- #
def construir() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo Maestro"

    # ---- Encabezados ----
    for idx, titulo in enumerate(ENCABEZADOS, start=1):
        celda = ws.cell(row=1, column=idx, value=titulo)
        letra = get_column_letter(idx)
        celda.fill = FORMULA_HEADER_FILL if letra in COLS_FORMULA else HEADER_FILL
        celda.font = HEADER_FONT
        celda.alignment = CENTER
        celda.border = BORDER

    # ---- Filas de datos (2..21) ----
    for i, (dpi, desc, marca, modelo, anio, compra, autozone) in enumerate(PRODUCTOS):
        r = i + 2  # primera fila de datos = 2

        # A DPI (texto, conserva ceros a la izquierda como "0050")
        c = ws.cell(row=r, column=1, value=str(dpi)); c.number_format = "@"
        ws.cell(row=r, column=2, value=desc)                       # B Descripción
        ws.cell(row=r, column=3, value=marca)                      # C Marca
        ws.cell(row=r, column=4, value=modelo)                     # D Modelo
        c = ws.cell(row=r, column=5, value=anio); c.number_format = "0"            # E Año
        c = ws.cell(row=r, column=6, value=compra); c.number_format = MXN          # F Precio Compra
        c = ws.cell(row=r, column=7, value=autozone); c.number_format = MXN        # G Precio AutoZone

        # H Precio Cliente Final (FÓRMULA)
        c = ws.cell(row=r, column=8,
                    value=(f"=IF(G{r}>0,"
                           f"ROUNDUP(IF(G{r}-F{r}-250<1200,F{r}+1200,G{r}-250)/50,0)*50,"
                           f"ROUNDUP((F{r}+1500)/50,0)*50)"))
        c.number_format = MXN

        # I Precio Taller (FÓRMULA)
        c = ws.cell(row=r, column=9,
                    value=f"=ROUNDUP(MAX(F{r}+800,H{r}*0.82)/50,0)*50")
        c.number_format = MXN

        # J Utilidad Cliente (FÓRMULA)
        c = ws.cell(row=r, column=10, value=f"=H{r}-F{r}"); c.number_format = MXN
        # K Utilidad Taller (FÓRMULA)
        c = ws.cell(row=r, column=11, value=f"=I{r}-F{r}"); c.number_format = MXN

        # L Stock Actual
        c = ws.cell(row=r, column=12, value=STOCK_ACTUAL[i]); c.number_format = "0"
        # M Stock Mínimo
        c = ws.cell(row=r, column=13, value=STOCK_MINIMO); c.number_format = "0"
        # N Stock Ideal
        c = ws.cell(row=r, column=14, value=STOCK_IDEAL); c.number_format = "0"

        # O Proveedor Principal
        ws.cell(row=r, column=15, value=PROVEEDOR)

        # P Fecha Última Compra
        c = ws.cell(row=r, column=16,
                    value=HOY - timedelta(days=DIAS_ULTIMA_COMPRA[i]))
        c.number_format = FECHA

        # Q Fecha Última Venta (puede quedar vacía -> "Sin venta")
        dv = DIAS_ULTIMA_VENTA[i]
        if dv is not None:
            c = ws.cell(row=r, column=17, value=HOY - timedelta(days=dv))
            c.number_format = FECHA

        # R Estado (dropdown)
        ws.cell(row=r, column=18, value="Activo")

        # S Rotación (FÓRMULA)
        ws.cell(row=r, column=19,
                value=(f'=IF(Q{r}="","Sin venta",'
                       f'IF(TODAY()-Q{r}<=90,"Estrella",'
                       f'IF(TODAY()-Q{r}<=180,"Media",'
                       f'IF(TODAY()-Q{r}<=365,"Lenta","Muerta"))))'))

        # T Alerta AutoZone (FÓRMULA)
        ws.cell(row=r, column=20,
                value=f'=IF(G{r}="","Pendiente",IF(G{r}>0,"OK","Revisar"))')

        # U Observaciones
        if i in OBSERVACIONES:
            ws.cell(row=r, column=21, value=OBSERVACIONES[i])

        # Bordes + alineación de toda la fila
        for col in range(1, 22):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER
            if col in (2, 15, 21):       # textos largos a la izquierda
                cell.alignment = LEFT
            else:
                cell.alignment = CENTER

    aplicar_validaciones(ws)
    aplicar_formato_condicional(ws)
    aplicar_diseno(ws)

    # Forzar recálculo al abrir (las fórmulas no traen valor en caché)
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    return wb


# --------------------------------------------------------------------------- #
# Validaciones de datos                                                       #
# --------------------------------------------------------------------------- #
def aplicar_validaciones(ws):
    # 1) DPI único (sin duplicados)
    dv_dpi = DataValidation(
        type="custom", formula1="COUNTIF($A$2:$A$1000,A2)=1",
        allow_blank=True, showErrorMessage=True,
        errorTitle="DPI duplicado",
        error="Ese DPI ya existe en el catálogo. El DPI debe ser único.")
    dv_dpi.add("A2:A1000")
    ws.add_data_validation(dv_dpi)

    # 2) Precio Compra > 0
    dv_compra = DataValidation(
        type="decimal", operator="greaterThan", formula1="0",
        allow_blank=True, showErrorMessage=True,
        errorTitle="Precio inválido",
        error="El Precio Compra debe ser un número mayor a 0.")
    dv_compra.add("F2:F1000")
    ws.add_data_validation(dv_compra)

    # 3,4,5) Stock Actual / Mínimo / Ideal >= 0
    for col in ("L", "M", "N"):
        dv = DataValidation(
            type="whole", operator="greaterThanOrEqual", formula1="0",
            allow_blank=True, showErrorMessage=True,
            errorTitle="Stock inválido",
            error="El stock no puede ser negativo (debe ser 0 o mayor).")
        dv.add(f"{col}2:{col}1000")
        ws.add_data_validation(dv)

    # 6) Estado: dropdown Activo / Inactivo
    dv_estado = DataValidation(
        type="list", formula1='"Activo,Inactivo"',
        allow_blank=True, showErrorMessage=True,
        showDropDown=False,  # False = SÍ muestra la flecha del desplegable
        errorTitle="Estado inválido",
        error="Selecciona Activo o Inactivo de la lista.")
    dv_estado.add("R2:R1000")
    ws.add_data_validation(dv_estado)


# --------------------------------------------------------------------------- #
# Formato condicional                                                         #
# --------------------------------------------------------------------------- #
def aplicar_formato_condicional(ws):
    # --- Stock Actual (L) ---
    rango_L = "L2:L1000"
    ws.conditional_formatting.add(rango_L, FormulaRule(formula=["$L2=0"], fill=F_ROJO, font=T_ROJO))
    ws.conditional_formatting.add(rango_L, FormulaRule(formula=["AND($L2>0,$L2<$M2)"], fill=F_NARAN, font=T_NARAN))
    ws.conditional_formatting.add(rango_L, FormulaRule(formula=["AND($L2>=$M2,$L2<$N2)"], fill=F_AMARI, font=T_AMARI))
    ws.conditional_formatting.add(rango_L, FormulaRule(formula=["$L2>=$N2"], fill=F_VERDE, font=T_VERDE))

    # --- Rotación (S) ---
    rango_S = "S2:S1000"
    ws.conditional_formatting.add(rango_S, CellIsRule(operator="equal", formula=['"Estrella"'], fill=F_VERDE, font=T_VERDE))
    ws.conditional_formatting.add(rango_S, CellIsRule(operator="equal", formula=['"Media"'], fill=F_AMARI, font=T_AMARI))
    ws.conditional_formatting.add(rango_S, CellIsRule(operator="equal", formula=['"Lenta"'], fill=F_NARAN, font=T_NARAN))
    ws.conditional_formatting.add(rango_S, CellIsRule(operator="equal", formula=['"Muerta"'], fill=F_ROJO, font=T_ROJO))
    ws.conditional_formatting.add(rango_S, CellIsRule(operator="equal", formula=['"Sin venta"'], fill=F_GRIS, font=T_GRIS))

    # --- Utilidad Cliente (J) ---
    rango_J = "J2:J1000"
    ws.conditional_formatting.add(rango_J, FormulaRule(formula=["$J2<1200"], fill=F_ROJO, font=T_ROJO))
    ws.conditional_formatting.add(rango_J, FormulaRule(formula=["AND($J2>=1200,$J2<1500)"], fill=F_AMARI, font=T_AMARI))
    ws.conditional_formatting.add(rango_J, FormulaRule(formula=["$J2>=1500"], fill=F_VERDE, font=T_VERDE))


# --------------------------------------------------------------------------- #
# Diseño (anchos, panel congelado, autofiltro)                                #
# --------------------------------------------------------------------------- #
def aplicar_diseno(ws):
    anchos = {
        "A": 12, "B": 28, "C": 15, "D": 12, "E": 7, "F": 14, "G": 15,
        "H": 18, "I": 14, "J": 16, "K": 15, "L": 11, "M": 12, "N": 11,
        "O": 22, "P": 16, "Q": 16, "R": 11, "S": 12, "T": 15, "U": 30,
    }
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "B2"                 # congela encabezado + columna DPI
    ws.auto_filter.ref = "A1:U21"          # autofiltro sobre los datos


# --------------------------------------------------------------------------- #
# Verificación (pruebas del Sprint 1)                                         #
# --------------------------------------------------------------------------- #
def verificar():
    print("\n" + "=" * 96)
    print("VERIFICACIÓN DE FÓRMULAS (cálculo replicado en Python)")
    print("=" * 96)
    print(f"{'DPI':<10}{'Compra':>8}{'AutoZ':>8}{'Cliente':>9}{'Taller':>8}"
          f"{'UtilCli':>9}{'UtilTal':>9}{'Rotación':>11}{'Alerta':>9}")
    print("-" * 96)

    errores = 0
    for i, (dpi, desc, marca, modelo, anio, compra, autozone) in enumerate(PRODUCTOS):
        cliente = precio_cliente(compra, autozone)
        taller = precio_taller(compra, cliente)
        util_cli = cliente - compra
        util_tal = taller - compra

        # Rotación
        dv = DIAS_ULTIMA_VENTA[i]
        if dv is None:
            rot = "Sin venta"
        elif dv <= 90:
            rot = "Estrella"
        elif dv <= 180:
            rot = "Media"
        elif dv <= 365:
            rot = "Lenta"
        else:
            rot = "Muerta"
        alerta = "OK" if autozone > 0 else "Pendiente"

        print(f"{dpi:<10}{compra:>8}{autozone:>8}{cliente:>9}{taller:>8}"
              f"{util_cli:>9}{util_tal:>9}{rot:>11}{alerta:>9}")

        # --- Aserciones (pruebas obligatorias del Sprint) ---
        if cliente % 50 != 0:
            print(f"   ✗ Cliente no múltiplo de 50: {cliente}"); errores += 1
        if taller % 50 != 0:
            print(f"   ✗ Taller no múltiplo de 50: {taller}"); errores += 1
        if util_cli < 1200:
            print(f"   ✗ Utilidad Cliente < 1200: {util_cli}"); errores += 1
        if util_tal < 800:
            print(f"   ✗ Utilidad Taller < 800: {util_tal}"); errores += 1

    print("-" * 96)
    # DPI únicos
    dpis = [p[0] for p in PRODUCTOS]
    if len(dpis) != len(set(dpis)):
        print("   ✗ Hay DPI duplicados"); errores += 1

    if errores == 0:
        print("✓ TODAS LAS PRUEBAS PASARON")
        print("  • Todos los precios terminan en 00 o 50")
        print("  • Utilidad Cliente >= $1,200 en todos los productos")
        print("  • Utilidad Taller  >= $800 en todos los productos")
        print("  • 20 DPI únicos, sin duplicados")
    else:
        print(f"✗ {errores} ERROR(ES) ENCONTRADO(S)")
    print("=" * 96 + "\n")
    return errores


# --------------------------------------------------------------------------- #
if __name__ == "__main__":
    import os

    salida = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                          "ENTREGAS", "SYSEA_Radiadores_SPRINT1.xlsx")
    errores = verificar()
    wb = construir()
    wb.save(salida)
    print(f"Archivo generado: {salida}")
    if errores:
        raise SystemExit(1)
